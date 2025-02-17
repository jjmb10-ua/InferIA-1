import os
import time
import pyautogui
import subprocess
from openpyxl import load_workbook
from pydantic import BaseModel
import openai
from typing import Optional
import base64
import json
import ast

# Configurar la API Key
client = openai.Client(api_key="")

# Open a file, take screenshot and save the image
def takeScreenshot(filename, output_dir="images"):
    # Verify if the file exists
    if not os.path.exists(filename):
        print(f"File does not exists: {filename}")
        return

    # Create the directory
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Load the Excel file to obtain the sheets
    try:
        wb = load_workbook(filename, read_only=True)
        sheets = wb.sheetnames
        print(f"Sheets founds: {sheets}")
    except Exception as e:
        print(f"Error processing sheets: {e}")
        return

    # Open the file
    try:
        print(f"Opening the file: {filename}. Waiting for Excel to fully load...")
        subprocess.Popen(['start', filename], shell=True)
        time.sleep(5)  # Wait for the file to open completely
    except Exception as e:
        print(f"Error opening the file: {e}")
        return

    # Put Excel in full screen
    try:
        pyautogui.hotkey("alt", "space")  # Open the window menu
        time.sleep(1)
        pyautogui.press("x")  # Maximize de window
        time.sleep(1)
        # Toggles twice
        pyautogui.hotkey("ctrl", "f1")
        time.sleep(1)
        pyautogui.hotkey("ctrl", "f1")
        time.sleep(1)
    except Exception as e:
        print(f"Error maximizing excel: {e}")
        return

    # Ensure that we are in the first sheet 
    try:
        print("Forzando ir a la primera hoja...")
        wb = load_workbook(filename, read_only=True)
        sheets = wb.sheetnames
        wb.close()

        for _ in range(len(sheets)):
            pyautogui.hotkey("ctrl", "pgup")
            time.sleep(1)
    except Exception as e:
        print(f"Error al forzar la primera hoja: {e}")

    # Take the screenshot of each sheet
    for i, sh in enumerate(sheets):
        try:
            # Change the sheet using (Ctrl+PgDown o Ctrl+PgUp)
            if i > 0:
                pyautogui.hotkey("ctrl", "pgdn")  # Navigate to the next sheet
                time.sleep(2)  # Wait for the sheet to load

            # Take the screenshot
            screenshot = pyautogui.screenshot()
            filename = f"sh{i + 1}_{sh}.png"
            filepath = os.path.join(output_dir, filename)
            screenshot.save(filepath)
            print(f"Screenshot saved in: {filepath}")
        except Exception as e:
            print(f"Error taking the screenshot for the sheet {sh}: {e}")

    # Close the file
    try:
        print("Closing the file...")
        pyautogui.hotkey("alt", "f4")
        print("File closed correctly.")
    except Exception as e:
        print(f"Error closing the file: {e}")
    finally:
        wb.close()


# Función para crear una clase dinámica de Pydantic
def createDynamicClass(class_name: str, attributes: list):
    # Usamos `__annotations__` para definir los tipos correctamente
    fields = {attr: Optional[str] for attr in attributes}  # Solo anotación de tipo
    defaults = {attr: None for attr in attributes}  # Valores predeterminados
    
    # Crear la clase con `type`
    dynamic_class = type(class_name, (BaseModel,), {"__annotations__": fields, **defaults})
    return dynamic_class

# Function to encode the image
def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode("utf-8")

# Convierte la respuesta en diccionario y la guarda como JSON
def save_dict_to_json(response_model, outputFile):
    try:
        # Convertir la respuesta del modelo Pydantic a un diccionario
        response_dict = response_model.model_dump()

        # Guardar el diccionario en un archivo JSON
        with open(outputFile, "w", encoding="utf-8") as json_file:
            json.dump(response_dict, json_file, indent=4, ensure_ascii=False)

        print(f"Datos guardados en {outputFile}")
    except Exception as e:
        print(f"Error al guardar los datos en JSON: {e}")

def processMetadata(fname, sh_index, sh, attributes, data_start):
    # Create dynamic class with specified attributes
    DynamicModel = createDynamicClass("DynamicModel", attributes)

    # Path to your image
    image_path = f"images/sh{sh_index}_{sh.title}.png"
    # Getting the base64 string
    base64_image = encode_image(image_path)

    completion = client.beta.chat.completions.parse(
        model="gpt-4o-2024-08-06",
        messages=[
            {"role": "system", "content": "You are an expert at structured data extraction. You will be given an image from an Excel sheet and should convert it into the given structure."},
            {"role": "user", "content": [
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/png;base64,{base64_image}"},
                    },
                ],
            }
        ],
        response_format=DynamicModel,
    )

    research_paper = completion.choices[0].message.parsed
    print("Response from the API:")
    print(research_paper)

    # Saved as json
    base_filename = os.path.basename(fname)
    base_filename_without_ext, _ = os.path.splitext(base_filename)

    output_dir = os.path.join(".", "json")
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    outputFile = os.path.join(output_dir, f"{base_filename_without_ext}_sh{sh_index}_{sh.title}.json")
    save_dict_to_json(research_paper, outputFile)

    usage = completion.usage
    if usage:
        total_tokens = usage.total_tokens
        print(f"Total tokens used: {total_tokens}")
    else:
        print("There is no information abaout usage.")

    # return research_paper.headers

##############################################
# HEADERS #
##############################################

def clean_response(response):
    """
    Elimina marcas de código Markdown y convierte la respuesta de GPT
    a una lista de Python (si es posible).
    """
    if response.startswith("```python"):
        response = response[len("```python"):].strip()
    if response.endswith("```"):
        response = response[:-len("```")].strip()
    
    try:
        parsed_response = ast.literal_eval(response)
        if isinstance(parsed_response, list):
            return parsed_response
        else:
            print("Error: La respuesta no es una lista.")
            return []
    except (SyntaxError, ValueError) as e:
        print(f"Error al procesar la respuesta como lista: {e}")
        return []
    
def findHeaders(sh_index, sh, previous_headers=None):
    if previous_headers is None:
        previous_headers = []

    # Excel image 
    image_path = f"images/sh{sh_index}_{sh.title}.png"
    base64_image = encode_image(image_path)

    # Preparamos el mensaje del sistema con instrucciones más precisas
    # acerca de conservar el idioma y, si falta un encabezado, inventarlo
    # en el idioma local.
    system_prompt = f"""
Eres un experto en extracción de datos. Debes devolver estrictamente una lista de Python con los encabezados 
tal como aparecen en el idioma de la hoja. 

- NO traduzcas ni modifiques los encabezados que sí aparezcan en la tabla.
- Si hay un encabezado que se subdivide en dos, el nombre del principal debe aparecer en ambos
- Si un encabezado no existe en la tabla, pero deduces que debería existir (comparando con datos o con encabezados anteriores),
  entonces inventa un nombre de encabezado en el idioma local que observas en esta hoja.
- A modo de referencia, en hojas anteriores se encontraron estos encabezados: {previous_headers}.
  Si alguno coincide conceptualmente pero está en distinto idioma, respeta el idioma de esta hoja.
- Devuelve SOLO la lista con los encabezados, sin texto adicional ni explicaciones.
"""

    completion = client.chat.completions.create(
        model="gpt-4o-2024-08-06",
        messages=[
            {"role": "system", "content": system_prompt.strip()},
            {
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/png;base64,{base64_image}"},
                    },
                ],
            }
        ],
    )

    research_paper = completion.choices[0].message.content
    cleaned_response = clean_response(research_paper)
    print("\nResponse from the API:")
    print(cleaned_response)

    usage = completion.usage
    if usage:
        total_tokens = usage.total_tokens
        print(f"Total tokens utilizados: {total_tokens}")
    else:
        print("No se pudo obtener información sobre el uso de tokens.")

    return cleaned_response

##############################################
# DATA #
##############################################

def decide_subgroup_column_name_via_gpt(subgroup_values):
    if not subgroup_values:
        return "Subgrupo"
    
    # Preparamos el prompt
    system_prompt = (
        "Eres un sistema que recibe una lista de subgrupos (texto) encontrados en una tabla. "
        "Tu tarea es proponer un único nombre de columna (en singular o plural) que mejor describa "
        "qué representan esos subgrupos en el idioma de los datos. Devuelve estrictamente un JSON con la forma:\n"
        "{\"column_name\": \"texto\"}\n"
        "Nada más."
    )
    user_prompt = f"Subgrupos encontrados: {subgroup_values}"

    completion = client.chat.completions.create(
        model="gpt-4o-2024-08-06",
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    )
    response = completion.choices[0].message.content.strip()

    # Trying to parse the response as dict with ast.literal_eval
    import ast
    try:
        parsed = ast.literal_eval(response)
        if isinstance(parsed, dict) and "column_name" in parsed:
            return parsed["column_name"]
        else:
            return "Subgrupo"
    except Exception:
        return "Subgrupo"