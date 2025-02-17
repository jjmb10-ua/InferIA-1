from openpyxl import load_workbook
import pandas as pd
import os
import metaData

# Finds the starting line of the data based on the most frequent number of non-empty columns
def findDataStart(sheet):
    # Dictionary to store statistics
    column_stats = {}

    # Iterate over the rows in the sheet
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # Count the number of non-empty cells in the row
        non_empty_cells = [cell for cell in row if cell is not None]

        # If this number of columns is not in the dictionary, initialize it
        if len(non_empty_cells) not in column_stats:
            
            column_stats[len(non_empty_cells)] = {
                "first_row": row_idx,  # Store the first row where it appears
                "count": 0  # Initial counter
            }
        # Increment the counter
        column_stats[len(non_empty_cells)]["count"] += 1

    # If no rows were found, return -1
    if not column_stats:
        return -1

    # Find the number of columns with the highest occurrences
    most_frequent = max(column_stats.items(), key=lambda x: x[1]["count"])
    most_frequent_row = most_frequent[1]["first_row"]

    # print(f"Column statistics: {column_stats}")
    return most_frequent_row

def fillHeader(headers, total_columns):
    cont = 1
    while len(headers) < total_columns:
        if cont == 1:
            headers = [cont] + headers
        else:
            headers = headers + [cont]
        cont += 1
    return headers

def isTotalRow(sheet, row_idx):
    for col_idx in range(1, sheet.max_column + 1):
        column_values = [
            sheet.cell(row=r, column=col_idx).value
            for r in range(1, sheet.max_row + 1)
            if r != row_idx and isinstance(sheet.cell(row=r, column=col_idx).value, (int, float))
        ]
        current_value = sheet.cell(row=row_idx, column=col_idx).value
        if isinstance(current_value, (int, float)) and current_value == sum(column_values):
            return True
    return False

def extractData(sh, headers, data_start_row):
    current_subgroup = None
    data = []
    found_subgroups = set()  # To save all the subgroups

    # Normalize expected headers to lowercase
    normalized_headers = [header.lower() for header in headers]

    # Read the first row of data and filter out empty cells
    first_data_row_cells = list(sh.iter_rows(min_row=data_start_row,
                                             max_row=data_start_row,
                                             values_only=True))[0]
    filtered_first = [cell for cell in first_data_row_cells if cell is not None]
    
    # If there are more columns than headers, call fillHeader to complete the headers
    if len(filtered_first) > len(headers):
        headers = fillHeader(headers, len(filtered_first))

    # Read all rows from the data_start_row
    for row_idx, row in enumerate(sh.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        # Filter out empty cells from the row
        filtered_row = [cell for cell in row if cell is not None]

        # Normalize the row cells to lowercase
        normalized_row = [str(cell).lower() for cell in filtered_row]

        # Avoid empty rows
        if not filtered_row:
            continue

        # Check if the row matches the expected headers, to skip it
        if set(normalized_row) == set(normalized_headers):
            # Skip this row because it appears to be a header
            continue
        # Skip if one row value matches with a header
        if any(cell in normalized_headers for cell in normalized_row):
            continue

        # Check if the row is a "total row" and discard it if so
        if isTotalRow(sh, row_idx):
            continue

        # If the row only has one row, it can be a subgroup
        # FUTURE IMPLEMENTHATION: ask model vision if is a subgroup
        if 1 == len(filtered_row):
            possible_subgroup = " ".join(str(x) for x in filtered_row).strip()
            current_subgroup = possible_subgroup
            found_subgroups.add(current_subgroup)
            continue

        # Add "Subgrupo" at the end of the headers
        if current_subgroup and "Subgrupo" not in headers:
            headers.append("Subgrupo")

        # Create a dictionary for each row, mapping headers to their respective column values
        row_dict = {}
        for i, h in enumerate(headers):
            if h == "Subgrupo":
                row_dict[h] = current_subgroup
            else:
                # Avoid out-of-range
                if i < len(filtered_row):
                    row_dict[h] = filtered_row[i]
                else:
                    row_dict[h] = None
                    
        data.append(row_dict)

    # Convert the data into a pandas DataFrame
    df = pd.DataFrame(data)
    return headers, df, found_subgroups

def save_to_csv(data, headers, output_file):
    if headers and isinstance(data, pd.DataFrame) and not data.empty:
        data = data.iloc[:, :len(headers)]
        data.columns = headers
        data.to_csv(output_file, index=False, encoding="utf-8", float_format='%.4f')
        print(f"Data saved in {output_file}")
    else:
        print("No data table found on sheet")

def processData(fname, sh_index, sh, headers, data_start_row):
    
    if data_start_row == -1:
        print("There are not data rows.")
        return
    else:
        print(f"Data begin in the row: {data_start_row}")

    # Extract data and detected subgroups
    final_headers, extracted_data, found_subgroups = extractData(sh, headers, data_start_row)

    # Call model to decide the real name of "Subgrupo"
    if found_subgroups:
        column_name = metaData.decide_subgroup_column_name_via_gpt(list(found_subgroups))
        print(f"GPT suggests calling column subgroups: {column_name}")

    # Rename the column in the DataFrame
    if "Subgrupo" in extracted_data.columns:
        extracted_data.rename(columns={"Subgrupo": column_name}, inplace=True)
        # Update name on final_headers
        final_headers = [column_name if h == "Subgrupo" else h for h in final_headers]

    # Saving in CSV
    base_filename = os.path.basename(fname)
    base_filename_without_ext, _ = os.path.splitext(base_filename)

    output_dir = os.path.join(".", "csv")
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    outputFile = os.path.join(output_dir, f"{base_filename_without_ext}_sh{sh_index}_{sh.title}.csv")
    save_to_csv(extracted_data, final_headers, outputFile)