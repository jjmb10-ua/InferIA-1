#Author: Juan José Martínez Berná
import os
import sys
import re
import json
import pandas as pd
from openpyxl import Workbook, load_workbook
import pyautogui
import subprocess
import time
import openai
from pydantic import BaseModel
from typing import Optional, List
import base64
import csv

# Configurar la API Key
client = openai.Client(api_key="")

# Open a file, take screenshot and save the image
def takeScreenshot(filename, output_dir="images"):
    # Verify if the file exists
    if not os.path.exists(filename):
        print(f"File does not exists: {filename}")
        return

    # Create the directory
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Load the Excel file to obtain the sheets
    try:
        wb = load_workbook(filename, read_only=True)
        sheets = wb.sheetnames
        print(f"Sheets founds: {sheets}")
    except Exception as e:
        print(f"Error processing sheets: {e}")
        return

    # Open the file
    try:
        print(f"Opening the file: {filename}. Waiting for Excel to fully load...")
        subprocess.Popen(['start', filename], shell=True)
        time.sleep(5)  # Wait for the file to open completely
    except Exception as e:
        print(f"Error opening the file: {e}")
        return

    # Put Excel in full screen
    try:
        pyautogui.hotkey("alt", "space")  # Open the window menu
        time.sleep(1)
        pyautogui.press("x")  # Maximize de window
        time.sleep(1)
        pyautogui.hotkey("ctrl", "f1")  # Hide options
        time.sleep(2)
    except Exception as e:
        print(f"Error maximizing excel: {e}")
        return

    # Take the screenshot of each sheet
    for i, sh in enumerate(sheets):
        try:
            # Change the sheet using (Ctrl+PgDown o Ctrl+PgUp)
            if i > 0:
                pyautogui.hotkey("ctrl", "pgdn")  # Navigate to the next sheet
                time.sleep(2)  # Wait for the sheet to load

            # Take the screenshot
            screenshot = pyautogui.screenshot()
            filename = f"sh{i + 1}_{sh}.png"
            filepath = os.path.join(output_dir, filename)
            screenshot.save(filepath)
            print(f"Screenshot saved in: {filepath}")
        except Exception as e:
            print(f"Error taking the screenshot for the sheet {sh}: {e}")

    # Close the file
    try:
        print("Closing the file...")
        pyautogui.hotkey("alt", "f4")
        print("File closed correctly.")
    except Exception as e:
        print(f"Error closing the file: {e}")
    finally:
        wb.close()

# Function to create a dinamic class
def createDynamicClass(class_name: str, attributes: list):
    # Using `__annotations__` to define types correctly
    fields = {attr: Optional[str] for attr in attributes}  # Only type annotation
    defaults = {attr: None for attr in attributes}  # Default values
    
    # Adding the attribute 'headers'
    fields["headers"] = Optional[List[str]]
    defaults["headers"] = None

    # Create class with `type`
    dynamic_class = type(class_name, (BaseModel,), {"__annotations__": fields, **defaults})
    return dynamic_class

# Function to encode the image
def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode("utf-8")

# Convierte la respuesta en diccionario y la guarda como JSON
def save_dict_to_json(response_model, outputFile):
    try:
        # Convertir la respuesta del modelo Pydantic a un diccionario
        response_dict = response_model.model_dump()

        # Guardar el diccionario en un archivo JSON
        with open(outputFile, "w", encoding="utf-8") as json_file:
            json.dump(response_dict, json_file, indent=4, ensure_ascii=False)

        print(f"Datos guardados en {outputFile}")
    except Exception as e:
        print(f"Error al guardar los datos en JSON: {e}")

def save_to_csv(data, headers, output_file, fname):
    if headers and isinstance(data, pd.DataFrame) and not data.empty:
        # Ensure all rows have the same length as headers
        data = data.iloc[:, :len(headers)]  # Truncate extra columns if any

        # Set column names
        data.columns = headers

        # Save DataFrame to CSV
        data.to_csv(output_file, index=False, encoding="utf-8")
        print(f"Data saved to {output_file}")
    else:
        # Handle cases with no data or headers
        print(f"No table data found in sheet")

def processMetadata(sh_index, sh, attributes):
    # Create dynamic class with specified attributes
    DynamicModel = createDynamicClass("DynamicModel", attributes)

    # Path to your image
    image_path = f"images/sh{sh_index}_{sh.title}.png"
    # Getting the base64 string
    base64_image = encode_image(image_path)

    completion = client.beta.chat.completions.parse(
        model="gpt-4o-2024-08-06",
        messages=[
            {"role": "system", "content": "You are an expert at structured data extraction. You will be given an image from an Excel sheet and should convert it into the given structure."},
            {"role": "user", "content": [
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/png;base64,{base64_image}"},
                    },
                ],
            }
        ],
        response_format=DynamicModel,
    )
    research_paper = completion.choices[0].message.parsed
    print("\nResponse from the API:")
    print(research_paper)
    outputFile = f"sh{sh_index}_{sh.title}_metada.json"
    save_dict_to_json(research_paper, outputFile)

    # Obtener el uso de tokens del objeto `completion`
    usage = completion.usage  # Acceder directamente a la propiedad `usage`
    if usage:
        total_tokens = usage.total_tokens  # Acceder al atributo directamente
        print(f"Total tokens utilizados: {total_tokens}")
    else:
        print("No se pudo obtener información sobre el uso de tokens.")

    return research_paper.headers

def isValidHeader(row_cleaned):
    # Comprobar si cada celda parece un encabezado válido
    return all(cell.isalpha() or cell.replace(" ", "").isalnum() for cell in row_cleaned)

def getMissingHeader(fname):
    # Remove the file extension
    base_name = os.path.splitext(fname)[0]
    # Remove numbers and non-alphabetic characters
    words = re.findall(r'[a-zA-Z]+', base_name)
    # Capitalize the words
    cleaned_title = " ".join(word.capitalize() for word in words)

    return cleaned_title

# Finds the starting line of the data based on the most frequent number of non-empty columns
def findDataStart(sheet, expected_headers):
    # Minimum number of expected columns (based on the provided headers)
    min_columns = len(expected_headers)

    # Dictionary to store statistics
    column_stats = {}

    # Normalize expected headers to lowercase
    normalized_headers = [header.lower() for header in expected_headers]

    # Iterate over the rows in the sheet
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # Count the number of non-empty cells in the row
        non_empty_cells = [cell for cell in row if cell is not None]

        # Normalize the row cells to lowercase
        normalized_row = [str(cell).lower() for cell in non_empty_cells]

        # Check if the row matches the expected headers, to skip it
        if set(normalized_row) == set(normalized_headers):
            # Skip this row because it appears to be a header
            continue

        # Only record rows with at least the minimum number of expected columns
        if len(non_empty_cells) >= min_columns:
            if len(non_empty_cells) not in column_stats:
                # If this number of columns is not in the dictionary, initialize it
                column_stats[len(non_empty_cells)] = {
                    "first_row": row_idx,  # Store the first row where it appears
                    "count": 0  # Initial counter
                }
            # Increment the counter
            column_stats[len(non_empty_cells)]["count"] += 1

    # If no rows were found, return -1
    if not column_stats:
        return -1

    # Find the number of columns with the highest occurrences
    most_frequent = max(column_stats.items(), key=lambda x: x[1]["count"])
    most_frequent_row = most_frequent[1]["first_row"]

    # print(f"Column statistics: {column_stats}")
    return most_frequent_row

def fillHeader(headers, total_columns):
    cont = 1
    # Add placeholder headers for missing columns
    while len(headers) < total_columns:
        if cont == 1:
            headers = [cont] + headers
        else:
            headers = headers + [cont]
            
    return headers

def isTotalRow(sheet, row_idx):
    # Iterate over each column in the row
    for col_idx in range(1, sheet.max_column + 1):
        column_values = [
            sheet.cell(row=r, column=col_idx).value
            for r in range(1, sheet.max_row + 1)
            if r != row_idx and isinstance(sheet.cell(row=r, column=col_idx).value, (int, float))
        ]

        current_value = sheet.cell(row=row_idx, column=col_idx).value

        # Check if the current cell is equal to the sum of other cells in the column
        if isinstance(current_value, (int, float)) and current_value == sum(column_values):
            return True
    return False

def extractData(sh, headers, data_start_row):
    # Read the first row of data and filter out empty cells
    first_data_row = [
        cell for cell in list(sh.iter_rows(min_row=data_start_row, max_row=data_start_row, values_only=True))[0]
        if cell is not None
    ]

    # If there are more columns than headers, call fillHeader to complete the headers
    if len(first_data_row) > len(headers):
        headers = fillHeader(headers, len(first_data_row))

    # Read all rows from the data_start_row
    data = []
    for row_idx, row in enumerate(sh.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        # Filter out empty cells from the row
        filtered_row = [cell for cell in row if cell is not None]

        # Check if the row is a "total row" and discard it if so
        if isTotalRow(sh, row_idx):
            continue

        # Create a dictionary for each row, mapping headers to their respective column values
        row_dict = {headers[i]: filtered_row[i] if i < len(filtered_row) else None for i in range(len(headers))}
        data.append(row_dict)

    # Convert the data into a pandas DataFrame
    df = pd.DataFrame(data)
    return headers, df

def processData(fname, sh_index, sh, headers):

    # Find the most representative row
    data_start_row = findDataStart(sh, headers)

    if data_start_row == -1:
        print("No rows with data were found.")
        return
    else:
        print(f"Data starts on row: {data_start_row}")

    # Extract the data
    headers, extracted_data = extractData(sh, headers, data_start_row)

    # Save the data to a CSV file
    outputFile = f"sh{sh_index}_{sh.title}_data.csv"
    save_to_csv(extracted_data, headers, outputFile, fname)

# Processes each sheet and distinguish between metadata and data
def processSheet(fname, sh_index, sh, attributes):
    # Separate metadata and data
    headers = processMetadata(sh_index, sh, attributes)

    processData(fname, sh_index, sh, headers)

# Try to open the file and iterate all sheets
def openFile(filename, attributes):
    try:
        wb = load_workbook(filename, read_only=True)

        for sh_index, sheet in enumerate(wb, start=1):
            processSheet(filename, sh_index, sheet, attributes)

    except FileNotFoundError:
        print(f"Error: File '{filename}' not found.")
    except Exception as e:
        print(f"Oops... error : {e}")

def main():
    if len(sys.argv) < 3:
        print("Usage: python Code.py <EXCEL_SHEET> <ATTRIBUTES...>")
        sys.exit(1)

    # First argument is Excel sheet
    excel_path = sys.argv[1]
    # The following arguments are the metadata attributes
    attributes = sys.argv[2:]

    # Capture the entire screen from all sheets after opening the file
    takeScreenshot(excel_path)

    openFile(excel_path, attributes)

######################### MAIN ##########################

if __name__ == "__main__":
    main()